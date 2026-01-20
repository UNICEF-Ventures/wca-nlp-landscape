#!/usr/bin/env python3
"""
Convert the African Language Grid Excel spreadsheet to YAML format.

This script extracts data from the "Public Inventory of African Languages and Resources.xlsx"
spreadsheet and converts it to YAML files, preserving embedded hyperlinks.

Output files (in same directory as spreadsheet):
  - all_africa.yaml: Main language data from AllAfrica sheet
  - country_grid.yaml: Country-to-language mapping from CountryGrid sheet
  - column_sources.yaml: Header row with source URLs for each column
"""

import argparse
from pathlib import Path
import yaml
import openpyxl


# Custom YAML representer for multiline strings
def str_representer(dumper, data):
    if '\n' in data:
        return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
    return dumper.represent_scalar('tag:yaml.org,2002:str', data)


yaml.add_representer(str, str_representer)


# Column name normalization mapping
COLUMN_NAMES = {
    'Code': 'iso_639_3',
    'Name (not definitive)': 'name',
    'Alternate Names': 'alternate_names',
    '🇫🇷Nom': 'name_french',
    'Countries': 'countries',
    'Resource Page': 'resource_page',
    'Note': 'note',
    'Glotto🪵': 'glottocode',
    '🚌Wiki': 'wikipedia',
    'Population': 'population',
    '10ˣ': 'population_order',
    '⚠️Danger': 'endangerment',
    'Official': 'official_status',
    'Commission Page': 'acalan_commission',
    'ISO 639-1': 'iso_639_1',
    'Lanfrica Probe': 'lanfrica',
    'OLAC': 'olac',
    '👵🏦': 'grambank',
    'WALS': 'wals',
    'ELAR': 'elar',
    '⏳ELP': 'elp',
    '🌍🆔': 'afrolid',
    'CLDR': 'cldr',
    'Fineweb2': 'fineweb2',
    'AfricArXiV': 'africarxiv',
    'AfriCLIRMatrix': 'africlirmatrix',
    'Google Translate': 'google_translate',
    'mBERT': 'mbert',
    'NLLB 200': 'nllb_200',
    'mBart 50': 'mbart_50',
    'M2M-100': 'm2m_100',
    'Mozilla Common Voice': 'common_voice',
    'MADLAD-400 docs': 'madlad_400_docs',
    'MADLAD-400 sentences': 'madlad_400_sentences',
    'Aya 101': 'aya_101',
    '🕸️Webonary': 'webonary',
    'Language Box': 'language_box',
    'Kencorpus': 'kencorpus',
    'Bible': 'bible',
    '⌨️': 'keyboard',
    'NaijaVoices': 'naija_voices',
    'Fleurs': 'fleurs',
    'AfriHate': 'afrihate',
    'CMU Wilderness': 'cmu_wilderness',
    'QCRI Educational Domain Corpus': 'qcri_edu_corpus',
    'Wordnet': 'wordnet',
    'MAFAND': 'mafand',
    'IncubaLM': 'incubalm',
    'Sunflower': 'sunflower',
    '𝐀𝐟𝐫𝐢𝐪𝐮𝐞𝐋𝐋𝐌': 'afriquellm',
    'Swivuriso': 'swivuriso',
    'Autogramm': 'autogramm',
    'UD 2.17': 'universal_dependencies',
    'AfriMMT-EA': 'afrimmt_ea',
    '🎓¹🌍': 'education_primary_africa',
    '🎓²🌍': 'education_secondary_africa',
    '🎓³🌍': 'education_tertiary_africa',
    '🎓³🌎': 'education_tertiary_abroad',
}


def normalize_column_name(name):
    """Convert column header to a normalized key name."""
    if name in COLUMN_NAMES:
        return COLUMN_NAMES[name]
    # Fallback: lowercase, replace spaces with underscores
    return name.lower().replace(' ', '_').replace('-', '_')


def get_cell_value_with_link(data_cell, link_cell):
    """
    Extract cell value and hyperlink, returning a dict if there's a link,
    or just the value if there isn't.
    """
    value = data_cell.value
    if value is None:
        return None

    # Convert to string and clean up
    if isinstance(value, (int, float)):
        value = str(value)
    else:
        value = str(value).strip()

    if not value:
        return None

    # Check for hyperlink
    link = None
    if link_cell.hyperlink and link_cell.hyperlink.target:
        link = link_cell.hyperlink.target

    if link:
        return {'text': value, 'url': link}
    return value


def extract_all_africa(ws_data, ws_links):
    """Extract data from AllAfrica sheet."""
    # Get headers from row 2
    headers = []
    header_sources = {}

    for col_idx, (data_cell, link_cell) in enumerate(zip(ws_data[2], ws_links[2]), 1):
        header_val = data_cell.value
        if header_val:
            normalized = normalize_column_name(header_val)
            headers.append((col_idx, header_val, normalized))

            # Get source URL from row 2 if present
            if link_cell.hyperlink and link_cell.hyperlink.target:
                header_sources[normalized] = {
                    'original_name': header_val,
                    'source_url': link_cell.hyperlink.target
                }

    # Extract data rows (starting from row 3)
    languages = []
    for row_num in range(3, ws_data.max_row + 1):
        data_row = list(ws_data[row_num])
        link_row = list(ws_links[row_num])

        # Skip if no ISO code (column A)
        iso_code = data_row[0].value
        if not iso_code:
            continue

        lang_data = {}
        for col_idx, original_name, normalized_name in headers:
            cell_idx = col_idx - 1  # 0-indexed
            if cell_idx < len(data_row):
                value = get_cell_value_with_link(data_row[cell_idx], link_row[cell_idx])
                if value is not None:
                    lang_data[normalized_name] = value

        if lang_data:
            languages.append(lang_data)

    return languages, header_sources


def extract_country_grid(ws_data, ws_links):
    """Extract data from CountryGrid sheet, including hyperlinks."""
    # Row 2 contains country names
    # Row 3 contains language counts
    # Rows 4+ contain language names (with hyperlinks to Wikipedia/ScriptSource)

    countries = {}

    # Get country names from row 2 (skip first column)
    country_row = list(ws_data[2])

    for col_idx, cell in enumerate(country_row):
        country_name = cell.value
        if not country_name or not isinstance(country_name, str):
            continue
        if country_name.startswith('The initial'):  # Skip the explanatory text
            continue
        if country_name in ('Total Unique',):  # Skip metadata columns
            continue

        # Collect languages for this country
        languages = []
        for row_num in range(4, ws_data.max_row + 1):
            data_cell = ws_data.cell(row=row_num, column=col_idx + 1)
            link_cell = ws_links.cell(row=row_num, column=col_idx + 1)

            value = get_cell_value_with_link(data_cell, link_cell)
            if value is not None:
                languages.append(value)

        if languages:
            countries[country_name] = languages

    return countries


def main():
    parser = argparse.ArgumentParser(
        description='Convert African Language Grid Excel to YAML'
    )
    parser.add_argument(
        '--input',
        type=Path,
        default=Path('Source data/African-language-grid/Public Inventory of African Languages and Resources.xlsx'),
        help='Path to the Excel file'
    )
    args = parser.parse_args()

    input_path = args.input
    if not input_path.is_absolute():
        input_path = Path.cwd() / input_path

    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        return 1

    output_dir = input_path.parent

    print(f"Loading workbook: {input_path}")

    # Load with data_only=True to get calculated values
    wb_data = openpyxl.load_workbook(input_path, data_only=True)
    # Load without data_only to get hyperlinks
    wb_links = openpyxl.load_workbook(input_path, data_only=False)

    # Extract AllAfrica data
    print("Extracting AllAfrica sheet...")
    ws_data = wb_data['AllAfrica']
    ws_links = wb_links['AllAfrica']
    languages, column_sources = extract_all_africa(ws_data, ws_links)
    print(f"  Found {len(languages)} languages")

    # Extract CountryGrid data
    print("Extracting CountryGrid sheet...")
    ws_country_data = wb_data['CountryGrid']
    ws_country_links = wb_links['CountryGrid']
    country_grid = extract_country_grid(ws_country_data, ws_country_links)
    print(f"  Found {len(country_grid)} countries")

    # Write YAML files
    all_africa_path = output_dir / 'all_africa.yaml'
    print(f"Writing {all_africa_path}")
    with open(all_africa_path, 'w', encoding='utf-8') as f:
        yaml.dump(
            {'languages': languages},
            f,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            width=120
        )

    country_grid_path = output_dir / 'country_grid.yaml'
    print(f"Writing {country_grid_path}")
    with open(country_grid_path, 'w', encoding='utf-8') as f:
        yaml.dump(
            {'countries': country_grid},
            f,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            width=120
        )

    column_sources_path = output_dir / 'column_sources.yaml'
    print(f"Writing {column_sources_path}")
    with open(column_sources_path, 'w', encoding='utf-8') as f:
        yaml.dump(
            {'columns': column_sources},
            f,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            width=120
        )

    print("\nDone!")
    return 0


if __name__ == '__main__':
    exit(main())
